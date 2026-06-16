# Design Ref: §8.5 Seed Data — programmatic fixture builders for
# sermon_review tests. Generates byte buffers at runtime so no binary blobs
# are committed.

from __future__ import annotations

from copy import deepcopy
from datetime import datetime, timezone
from io import BytesIO

from openpyxl import Workbook, load_workbook

from app.sermon_review import (
    Segment,
    Sermon,
    build_xlsx,
)
from app.sermon_review.validation import COLUMNS


def make_golden_sermon() -> Sermon:
    now = datetime(2026, 6, 16, tzinfo=timezone.utc)
    segments = [
        Segment(
            segmentId="S001",
            order=1,
            original="오늘 우리는 하나님의 은혜를 보려고 합니다.",
            appTranslation="Today we will look at God's grace.",
            reviewedTranslation="Today we will look at God's grace.",
            notes="",
            status="Draft",
        ),
        Segment(
            segmentId="S002",
            order=2,
            original="은혜는 단지 좋은 감정이 아닙니다.",
            appTranslation="Grace is not merely a good feeling.",
            reviewedTranslation="Grace is not merely a good feeling.",
            notes="",
            status="Draft",
        ),
        Segment(
            segmentId="S003",
            order=3,
            original="그것은 하나님께서 주시는 선물입니다.",
            appTranslation="It is a gift given by God.",
            reviewedTranslation="It is a gift given by God.",
            notes="",
            status="Draft",
        ),
        Segment(
            segmentId="S004",
            order=4,
            original="우리는 이 은혜로 구원을 받습니다.",
            appTranslation="We are saved by this grace.",
            reviewedTranslation="We are saved by this grace.",
            notes="",
            status="Draft",
        ),
        Segment(
            segmentId="S005",
            order=5,
            original="기도합시다.",
            appTranslation="Let us pray.",
            reviewedTranslation="Let us pray.",
            notes="",
            status="Draft",
        ),
    ]
    return Sermon(
        sermonId="srm_test_abc123",
        orgId="org_test_xyz",
        title="Easter Sermon — Test Fixture",
        sourceType="paste",
        sourceRef=None,
        segments=segments,
        createdBy="uid_test_owner",
        createdAt=now,
        updatedAt=now,
    )


def valid_unmodified_xlsx(sermon: Sermon) -> bytes:
    return build_xlsx(sermon)


def valid_with_edits_xlsx(sermon: Sermon) -> bytes:
    data = build_xlsx(sermon)
    return _modify(
        data,
        edits={
            ("S001", "Reviewed Translation"):
                "Today, we will look together at the grace of God.",
            ("S001", "Status"): "Reviewed",
            ("S002", "Reviewed Translation"):
                "Grace is not just a comforting emotion.",
            ("S002", "Status"): "Reviewed",
            ("S003", "Notes"): "double-check translation",
        },
    )


def bad_missing_row_xlsx(sermon: Sermon) -> bytes:
    data = build_xlsx(sermon)
    return _delete_rows(data, segment_ids_to_delete={"S003"})


def bad_duplicate_xlsx(sermon: Sermon) -> bytes:
    data = build_xlsx(sermon)
    return _duplicate_row(data, segment_id_to_duplicate="S002")


def bad_wrong_sermon_xlsx(sermon: Sermon) -> bytes:
    data = build_xlsx(sermon)
    return _modify_column(
        data, column_name="Sermon ID", new_value="srm_different_other"
    )


def bad_mutated_original_xlsx(sermon: Sermon) -> bytes:
    data = build_xlsx(sermon)
    return _modify(
        data,
        edits={
            ("S003", "Original Text"): "이건 사용자가 실수로 수정한 원문입니다.",
        },
    )


def bad_missing_column_xlsx(sermon: Sermon) -> bytes:
    """Build a workbook missing the 'Reviewed Translation' column."""
    sermon_copy = deepcopy(sermon)
    wb = Workbook()
    ws = wb.active
    ws.title = "Sermon Review"
    cols = [c for c in COLUMNS if c != "Reviewed Translation"]
    for col_idx, name in enumerate(cols, start=1):
        ws.cell(row=1, column=col_idx, value=name)
    for row_idx, seg in enumerate(sermon_copy.segments, start=2):
        values: dict[str, object] = {
            "Sermon ID": sermon_copy.sermonId,
            "Segment ID": seg.segmentId,
            "Segment Order": seg.order,
            "Original Text": seg.original,
            "App Translation": seg.appTranslation,
            "Notes": seg.notes,
            "Status": seg.status,
        }
        for col_idx, name in enumerate(cols, start=1):
            ws.cell(row=row_idx, column=col_idx, value=values[name])
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def bad_csv_bytes() -> bytes:
    return b"Sermon ID,Segment ID,Original Text\nsrm_test_abc123,S001,foo\n"


def _column_index_by_name(headers: list[str], name: str) -> int | None:
    try:
        return headers.index(name)
    except ValueError:
        return None


def _modify(data: bytes, edits: dict[tuple[str, str], str]) -> bytes:
    wb = load_workbook(filename=BytesIO(data))
    ws = wb.active
    header_row = [cell.value for cell in ws[1]]
    headers = [str(h) if h else "" for h in header_row]
    segment_id_col = _column_index_by_name(headers, "Segment ID")
    if segment_id_col is None:
        raise RuntimeError("Test fixture broken: 'Segment ID' missing")

    for row in ws.iter_rows(min_row=2):
        sid = row[segment_id_col].value
        for (target_sid, target_col), new_val in edits.items():
            if sid == target_sid:
                col_idx = _column_index_by_name(headers, target_col)
                if col_idx is None:
                    raise RuntimeError(
                        f"Test fixture broken: column {target_col!r} missing"
                    )
                row[col_idx].value = new_val

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _modify_column(data: bytes, column_name: str, new_value: str) -> bytes:
    wb = load_workbook(filename=BytesIO(data))
    ws = wb.active
    header_row = [cell.value for cell in ws[1]]
    headers = [str(h) if h else "" for h in header_row]
    col_idx = _column_index_by_name(headers, column_name)
    if col_idx is None:
        raise RuntimeError(f"Test fixture broken: column {column_name!r} missing")
    for row in ws.iter_rows(min_row=2):
        row[col_idx].value = new_value
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _delete_rows(data: bytes, segment_ids_to_delete: set[str]) -> bytes:
    wb = load_workbook(filename=BytesIO(data))
    ws = wb.active
    header_row = [cell.value for cell in ws[1]]
    headers = [str(h) if h else "" for h in header_row]
    segment_id_col = _column_index_by_name(headers, "Segment ID")
    if segment_id_col is None:
        raise RuntimeError("Test fixture broken: 'Segment ID' missing")

    rows_to_delete: list[int] = []
    for row in ws.iter_rows(min_row=2):
        if row[segment_id_col].value in segment_ids_to_delete:
            rows_to_delete.append(row[0].row)

    for r in sorted(rows_to_delete, reverse=True):
        ws.delete_rows(r, 1)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _duplicate_row(data: bytes, segment_id_to_duplicate: str) -> bytes:
    wb = load_workbook(filename=BytesIO(data))
    ws = wb.active
    header_row = [cell.value for cell in ws[1]]
    headers = [str(h) if h else "" for h in header_row]
    segment_id_col = _column_index_by_name(headers, "Segment ID")
    if segment_id_col is None:
        raise RuntimeError("Test fixture broken: 'Segment ID' missing")

    source_values: list[object] | None = None
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[segment_id_col] == segment_id_to_duplicate:
            source_values = list(row)
            break
    if source_values is None:
        raise RuntimeError(
            f"Test fixture broken: segment {segment_id_to_duplicate!r} not found"
        )

    ws.append(source_values)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
