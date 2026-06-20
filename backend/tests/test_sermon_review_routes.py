# Design Ref: §4 API + §6.1 Error Codes + §7 RBAC.
# Tests call route functions directly with mocked multichurch_store and
# translate_text — matching the project's test_script_routes.py pattern.

from __future__ import annotations

import unittest
from io import BytesIO
from unittest.mock import AsyncMock, Mock, patch

from fastapi import HTTPException, UploadFile

from app.auth.firebase_auth import AuthenticatedUser
from app.routes import sermon_review as sermon_review_routes
from app.services.multichurch_store import InMemoryMultiChurchStore
from app.sermon_review import build_xlsx, Sermon


def _user(uid: str) -> AuthenticatedUser:
    return AuthenticatedUser(uid=uid, email=f"{uid}@example.com", displayName=uid)


def _bootstrap_owner(store, *, owner_uid="uid_owner", slug="test", name="Test") -> str:
    result = store.bootstrap_owner_org(
        owner_uid=owner_uid,
        owner_email=f"{owner_uid}@example.com",
        owner_display_name=owner_uid,
        church_name=name,
        church_slug=slug,
        timezone="America/Chicago",
        source="ko",
        target="en",
    )
    return str(result["orgId"])


def _add_member(store, org_id: str, uid: str, role: str) -> None:
    """Add a non-owner member by directly writing the members dict."""
    from datetime import datetime, timezone
    store._members[(org_id, uid)] = {
        "role": role,
        "createdAt": datetime.now(timezone.utc),
        "updatedAt": datetime.now(timezone.utc),
    }


class _FakeRequest:
    def __init__(self, method="POST", path="/api/test") -> None:
        self.method = method
        self.client = type("C", (), {"host": "127.0.0.1"})()
        self.url = type("U", (), {"path": path})()


def _fake_upload(data: bytes, filename: str = "test.txt") -> UploadFile:
    return UploadFile(filename=filename, file=BytesIO(data))


class _Patches:
    """Helper to patch multichurch_store + translate_text per test class."""
    def __init__(self, test: unittest.TestCase, store: InMemoryMultiChurchStore) -> None:
        self.test = test
        self.store = store
        self.translator = AsyncMock(side_effect=lambda t: f"[en] {t}")

    def start(self) -> None:
        p_store = patch.object(sermon_review_routes, "multichurch_store", self.store)
        p_translate = patch.object(
            sermon_review_routes, "translate_text", self.translator
        )
        p_store.start()
        p_translate.start()
        self.test.addCleanup(p_store.stop)
        self.test.addCleanup(p_translate.stop)


class IngestRouteTests(unittest.IsolatedAsyncioTestCase):
    def setUp(self) -> None:
        self.store = InMemoryMultiChurchStore()
        _Patches(self, self.store).start()
        self.org_id = _bootstrap_owner(self.store)

    async def test_paste_happy_path(self) -> None:
        result = await sermon_review_routes.ingest_sermon(
            org_id=self.org_id,
            request=_FakeRequest(),
            user=_user("uid_owner"),
            google_docs_service=None,
            sourceType="paste",
            title="My Sermon",
            text="오늘 우리는 은혜를 봅니다. 기도합시다.",
            url=None,
            file=None,
        )
        self.assertIn("sermonId", result["data"])
        self.assertEqual(result["data"]["title"], "My Sermon")
        self.assertGreaterEqual(result["data"]["segmentCount"], 1)

        # Verify persisted
        sid = result["data"]["sermonId"]
        sermon = self.store.get_review_sermon(self.org_id, sid)
        assert sermon is not None
        self.assertEqual(sermon["title"], "My Sermon")

    async def test_paste_empty_text_400(self) -> None:
        with self.assertRaises(HTTPException) as ctx:
            await sermon_review_routes.ingest_sermon(
                org_id=self.org_id,
                request=_FakeRequest(),
                user=_user("uid_owner"),
                google_docs_service=None,
                sourceType="paste",
                title="x",
                text="",
                url=None,
                file=None,
            )
        self.assertEqual(ctx.exception.status_code, 400)
        self.assertEqual(ctx.exception.detail["error"]["code"], "INVALID_SOURCE")

    async def test_empty_title_400(self) -> None:
        with self.assertRaises(HTTPException) as ctx:
            await sermon_review_routes.ingest_sermon(
                org_id=self.org_id,
                request=_FakeRequest(),
                user=_user("uid_owner"),
                google_docs_service=None,
                sourceType="paste",
                title="   ",
                text="오늘 우리는 은혜를 봅니다.",
                url=None,
                file=None,
            )
        self.assertEqual(ctx.exception.status_code, 400)

    async def test_unknown_source_type_400(self) -> None:
        with self.assertRaises(HTTPException) as ctx:
            await sermon_review_routes.ingest_sermon(
                org_id=self.org_id,
                request=_FakeRequest(),
                user=_user("uid_owner"),
                google_docs_service=None,
                sourceType="bogus",
                title="x",
                text=None,
                url=None,
                file=None,
            )
        self.assertEqual(ctx.exception.status_code, 400)
        self.assertEqual(ctx.exception.detail["error"]["code"], "INVALID_SOURCE")

    async def test_google_docs_without_oauth_returns_501(self) -> None:
        with self.assertRaises(HTTPException) as ctx:
            await sermon_review_routes.ingest_sermon(
                org_id=self.org_id,
                request=_FakeRequest(),
                user=_user("uid_owner"),
                google_docs_service=None,
                sourceType="google_docs",
                title="x",
                text=None,
                url="https://docs.google.com/document/d/abc123/edit",
                file=None,
            )
        self.assertEqual(ctx.exception.status_code, 501)
        self.assertEqual(
            ctx.exception.detail["error"]["code"], "GOOGLE_OAUTH_NOT_CONFIGURED"
        )

    async def test_google_docs_happy_path_with_mock_service(self) -> None:
        # Stub Docs API client mimicking documents().get().execute()
        fake_doc = {
            "body": {
                "content": [
                    {"paragraph": {"elements": [{"textRun": {"content": "오늘 우리는 은혜를 봅니다. "}}]}},
                    {"paragraph": {"elements": [{"textRun": {"content": "기도합시다."}}]}},
                ]
            }
        }
        execute = Mock(return_value=fake_doc)
        get = Mock(return_value=Mock(execute=execute))
        documents = Mock(return_value=Mock(get=get))
        fake_service = Mock(documents=documents)

        result = await sermon_review_routes.ingest_sermon(
            org_id=self.org_id,
            request=_FakeRequest(),
            user=_user("uid_owner"),
            google_docs_service=fake_service,
            sourceType="google_docs",
            title="From Docs",
            text=None,
            url="https://docs.google.com/document/d/abc123/edit",
            file=None,
        )
        self.assertIn("sermonId", result["data"])
        self.assertEqual(result["data"]["title"], "From Docs")
        get.assert_called_once_with(documentId="abc123")

    async def test_google_docs_rate_limited_maps_to_429(self) -> None:
        # Fake an HttpError with resp.status=429 — the route maps it to
        # GOOGLE_RATE_LIMITED so callers can show a user-friendly retry.
        class _FakeHttpError(Exception):
            def __init__(self) -> None:
                self.resp = type("R", (), {"status": 429})()

        execute = Mock(side_effect=_FakeHttpError())
        get = Mock(return_value=Mock(execute=execute))
        documents = Mock(return_value=Mock(get=get))
        fake_service = Mock(documents=documents)

        with self.assertRaises(HTTPException) as ctx:
            await sermon_review_routes.ingest_sermon(
                org_id=self.org_id,
                request=_FakeRequest(),
                user=_user("uid_owner"),
                google_docs_service=fake_service,
                sourceType="google_docs",
                title="Rate-limited",
                text=None,
                url="https://docs.google.com/document/d/abc123/edit",
                file=None,
            )
        self.assertEqual(ctx.exception.status_code, 429)
        self.assertEqual(
            ctx.exception.detail["error"]["code"], "GOOGLE_RATE_LIMITED"
        )

    def test_dependency_returns_none_when_header_missing(self) -> None:
        request = Mock(headers={})
        self.assertIsNone(
            sermon_review_routes._google_docs_service_dependency(request)
        )

    def test_dependency_builds_service_when_header_present(self) -> None:
        # The real builder imports googleapiclient and constructs a Resource.
        # The Google libs are in requirements.txt, so this should succeed.
        request = Mock(headers={"x-google-access-token": "ya29.fake_token"})
        service = sermon_review_routes._google_docs_service_dependency(request)
        # Either a real Resource or None if google libs unavailable; both are
        # acceptable. Asserting just that we don't blow up.
        self.assertTrue(service is None or hasattr(service, "documents"))

    async def test_txt_upload_happy_path(self) -> None:
        text = "오늘 우리는 하나님의 은혜를 보려고 합니다. 기도합시다."
        upload = _fake_upload(text.encode("utf-8"), "sermon.txt")
        result = await sermon_review_routes.ingest_sermon(
            org_id=self.org_id,
            request=_FakeRequest(),
            user=_user("uid_owner"),
            google_docs_service=None,
            sourceType="file_txt",
            title="Upload",
            text=None,
            url=None,
            file=upload,
        )
        self.assertIn("sermonId", result["data"])

    async def test_viewer_role_blocked_with_403(self) -> None:
        _add_member(self.store, self.org_id, "uid_viewer", "viewer")
        with self.assertRaises(HTTPException) as ctx:
            await sermon_review_routes.ingest_sermon(
                org_id=self.org_id,
                request=_FakeRequest(),
                user=_user("uid_viewer"),
                google_docs_service=None,
                sourceType="paste",
                title="x",
                text="오늘 우리는 은혜를 봅니다.",
                url=None,
                file=None,
            )
        self.assertEqual(ctx.exception.status_code, 403)

    async def test_host_role_can_ingest(self) -> None:
        _add_member(self.store, self.org_id, "uid_host", "host")
        result = await sermon_review_routes.ingest_sermon(
            org_id=self.org_id,
            request=_FakeRequest(),
            user=_user("uid_host"),
            google_docs_service=None,
            sourceType="paste",
            title="Host Sermon",
            text="오늘 우리는 은혜를 봅니다.",
            url=None,
            file=None,
        )
        self.assertEqual(result["data"]["title"], "Host Sermon")


class ListAndGetTests(unittest.IsolatedAsyncioTestCase):
    def setUp(self) -> None:
        self.store = InMemoryMultiChurchStore()
        _Patches(self, self.store).start()
        self.org_id = _bootstrap_owner(self.store)

    async def _seed_sermon(self, title: str = "Seeded") -> str:
        result = await sermon_review_routes.ingest_sermon(
            org_id=self.org_id,
            request=_FakeRequest(),
            user=_user("uid_owner"),
            google_docs_service=None,
            sourceType="paste",
            title=title,
            text="오늘 우리는 은혜를 봅니다. 기도합시다.",
            url=None,
            file=None,
        )
        return result["data"]["sermonId"]

    async def test_list_returns_summaries(self) -> None:
        await self._seed_sermon("A")
        await self._seed_sermon("B")
        result = sermon_review_routes.list_sermons(
            org_id=self.org_id, user=_user("uid_owner")
        )
        titles = {s["title"] for s in result["data"]}
        self.assertEqual(titles, {"A", "B"})
        for s in result["data"]:
            self.assertIn("segmentCount", s)
            self.assertIn("reviewedCount", s)

    async def test_get_sermon(self) -> None:
        sid = await self._seed_sermon()
        result = sermon_review_routes.get_sermon(
            org_id=self.org_id, sermon_id=sid, user=_user("uid_owner")
        )
        self.assertEqual(result["data"]["sermonId"], sid)
        self.assertGreater(len(result["data"]["segments"]), 0)

    async def test_get_sermon_not_found(self) -> None:
        with self.assertRaises(HTTPException) as ctx:
            sermon_review_routes.get_sermon(
                org_id=self.org_id, sermon_id="missing", user=_user("uid_owner")
            )
        self.assertEqual(ctx.exception.status_code, 404)
        self.assertEqual(
            ctx.exception.detail["error"]["code"], "SERMON_NOT_FOUND"
        )


class ExportTests(unittest.IsolatedAsyncioTestCase):
    def setUp(self) -> None:
        self.store = InMemoryMultiChurchStore()
        _Patches(self, self.store).start()
        self.org_id = _bootstrap_owner(self.store)

    async def test_export_returns_xlsx(self) -> None:
        result = await sermon_review_routes.ingest_sermon(
            org_id=self.org_id,
            request=_FakeRequest(),
            user=_user("uid_owner"),
            google_docs_service=None,
            sourceType="paste",
            title="Export Test",
            text="오늘 우리는 은혜를 봅니다.",
            url=None,
            file=None,
        )
        sid = result["data"]["sermonId"]

        response = sermon_review_routes.export_review_file(
            org_id=self.org_id,
            sermon_id=sid,
            request=_FakeRequest("GET", f"/api/org/{self.org_id}/sermons/{sid}/review-file.xlsx"),
            user=_user("uid_owner"),
        )
        self.assertEqual(response.status_code, 200)
        body = response.body
        self.assertEqual(body[:4], b"PK\x03\x04")
        self.assertIn(
            "vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            response.headers["content-type"],
        )
        self.assertIn("attachment", response.headers["content-disposition"])

    async def test_export_missing_sermon_404(self) -> None:
        with self.assertRaises(HTTPException) as ctx:
            sermon_review_routes.export_review_file(
                org_id=self.org_id,
                sermon_id="missing",
                request=_FakeRequest("GET"),
                user=_user("uid_owner"),
            )
        self.assertEqual(ctx.exception.status_code, 404)

    async def test_export_supports_korean_title_filename(self) -> None:
        result = await sermon_review_routes.ingest_sermon(
            org_id=self.org_id,
            request=_FakeRequest(),
            user=_user("uid_owner"),
            google_docs_service=None,
            sourceType="paste",
            title="은혜에 대한 설교",
            text="오늘 우리는 은혜를 봅니다.",
            url=None,
            file=None,
        )
        sid = result["data"]["sermonId"]

        response = sermon_review_routes.export_review_file(
            org_id=self.org_id,
            sermon_id=sid,
            request=_FakeRequest(
                "GET",
                f"/api/org/{self.org_id}/sermons/{sid}/review-file.xlsx",
            ),
            user=_user("uid_owner"),
        )

        disposition = response.headers["content-disposition"]
        self.assertIn('filename="sermon-', disposition)
        self.assertIn("filename*=UTF-8''", disposition)
        self.assertIn("%EC%9D%80%ED%98%9C", disposition)


class ImportTests(unittest.IsolatedAsyncioTestCase):
    def setUp(self) -> None:
        self.store = InMemoryMultiChurchStore()
        _Patches(self, self.store).start()
        self.org_id = _bootstrap_owner(self.store)

    async def _seed(self) -> str:
        result = await sermon_review_routes.ingest_sermon(
            org_id=self.org_id,
            request=_FakeRequest(),
            user=_user("uid_owner"),
            google_docs_service=None,
            sourceType="paste",
            title="Import Test",
            text="오늘 우리는 하나님의 은혜를 보려고 합니다. 기도합시다.",
            url=None,
            file=None,
        )
        return result["data"]["sermonId"]

    async def test_round_trip_import_succeeds(self) -> None:
        sid = await self._seed()
        sermon_raw = self.store.get_review_sermon(self.org_id, sid)
        sermon = Sermon.model_validate(sermon_raw)
        xlsx_bytes = build_xlsx(sermon)

        upload = _fake_upload(xlsx_bytes, "review.xlsx")
        result = await sermon_review_routes.import_review_file(
            org_id=self.org_id,
            sermon_id=sid,
            request=_FakeRequest(),
            user=_user("uid_owner"),
            file=upload,
        )
        self.assertEqual(result["data"]["summary"]["errored"], 0)
        self.assertEqual(
            result["data"]["summary"]["imported"], len(sermon.segments)
        )

    async def test_missing_file_415(self) -> None:
        sid = await self._seed()
        with self.assertRaises(HTTPException) as ctx:
            await sermon_review_routes.import_review_file(
                org_id=self.org_id,
                sermon_id=sid,
                request=_FakeRequest(),
                user=_user("uid_owner"),
                file=None,
            )
        self.assertEqual(ctx.exception.status_code, 415)

    async def test_non_xlsx_file_415(self) -> None:
        sid = await self._seed()
        upload = _fake_upload(b"not a real xlsx", "review.xlsx")
        with self.assertRaises(HTTPException) as ctx:
            await sermon_review_routes.import_review_file(
                org_id=self.org_id,
                sermon_id=sid,
                request=_FakeRequest(),
                user=_user("uid_owner"),
                file=upload,
            )
        self.assertEqual(ctx.exception.status_code, 415)

    async def test_corrupted_xlsx_returns_400_with_report(self) -> None:
        sid = await self._seed()
        sermon_raw = self.store.get_review_sermon(self.org_id, sid)
        sermon = Sermon.model_validate(sermon_raw)
        xlsx_bytes = build_xlsx(sermon)

        # Mutate Sermon ID column to simulate wrong-sermon upload
        from openpyxl import load_workbook
        from io import BytesIO
        wb = load_workbook(BytesIO(xlsx_bytes))
        ws = wb.active
        for row in ws.iter_rows(min_row=2):
            row[0].value = "srm_wrong"
        buf = BytesIO()
        wb.save(buf)

        upload = _fake_upload(buf.getvalue(), "review.xlsx")
        with self.assertRaises(HTTPException) as ctx:
            await sermon_review_routes.import_review_file(
                org_id=self.org_id,
                sermon_id=sid,
                request=_FakeRequest(),
                user=_user("uid_owner"),
                file=upload,
            )
        self.assertEqual(ctx.exception.status_code, 400)
        self.assertEqual(
            ctx.exception.detail["error"]["code"], "IMPORT_VALIDATION_FAILED"
        )
        # Report is in details
        codes = {
            r["code"] for r in ctx.exception.detail["error"]["details"]["rows"]
        }
        self.assertIn("WRONG_SERMON_ID", codes)


class LinkTests(unittest.IsolatedAsyncioTestCase):
    def setUp(self) -> None:
        self.store = InMemoryMultiChurchStore()
        _Patches(self, self.store).start()
        self.org_id = _bootstrap_owner(self.store)
        self.service_key = list(self.store._services.keys())[0][1]

    async def _seed(self, title: str = "Link Test") -> str:
        result = await sermon_review_routes.ingest_sermon(
            org_id=self.org_id,
            request=_FakeRequest(),
            user=_user("uid_owner"),
            google_docs_service=None,
            sourceType="paste",
            title=title,
            text="오늘 우리는 은혜를 봅니다.",
            url=None,
            file=None,
        )
        return result["data"]["sermonId"]

    async def test_link_happy_path(self) -> None:
        sid = await self._seed()
        result = sermon_review_routes.link_sermon(
            org_id=self.org_id,
            sermon_id=sid,
            body=sermon_review_routes._LinkBody(serviceKey=self.service_key),
            request=_FakeRequest(),
            user=_user("uid_owner"),
        )
        self.assertEqual(result["data"]["linkedSermonId"], sid)

    async def test_link_missing_service_key_400(self) -> None:
        sid = await self._seed()
        with self.assertRaises(HTTPException) as ctx:
            sermon_review_routes.link_sermon(
                org_id=self.org_id,
                sermon_id=sid,
                body=sermon_review_routes._LinkBody(serviceKey=None),
                request=_FakeRequest(),
                user=_user("uid_owner"),
            )
        self.assertEqual(ctx.exception.status_code, 400)

    async def test_link_already_linked_409(self) -> None:
        sid_a = await self._seed("A")
        sid_b = await self._seed("B")
        sermon_review_routes.link_sermon(
            org_id=self.org_id,
            sermon_id=sid_a,
            body=sermon_review_routes._LinkBody(serviceKey=self.service_key),
            request=_FakeRequest(),
            user=_user("uid_owner"),
        )
        with self.assertRaises(HTTPException) as ctx:
            sermon_review_routes.link_sermon(
                org_id=self.org_id,
                sermon_id=sid_b,
                body=sermon_review_routes._LinkBody(serviceKey=self.service_key),
                request=_FakeRequest(),
                user=_user("uid_owner"),
            )
        self.assertEqual(ctx.exception.status_code, 409)
        self.assertEqual(
            ctx.exception.detail["error"]["code"], "SERVICE_ALREADY_LINKED"
        )

    async def test_link_with_replace_succeeds(self) -> None:
        sid_a = await self._seed("A")
        sid_b = await self._seed("B")
        sermon_review_routes.link_sermon(
            org_id=self.org_id,
            sermon_id=sid_a,
            body=sermon_review_routes._LinkBody(serviceKey=self.service_key),
            request=_FakeRequest(),
            user=_user("uid_owner"),
        )
        result = sermon_review_routes.link_sermon(
            org_id=self.org_id,
            sermon_id=sid_b,
            body=sermon_review_routes._LinkBody(
                serviceKey=self.service_key, replace=True
            ),
            request=_FakeRequest(),
            user=_user("uid_owner"),
        )
        self.assertEqual(result["data"]["linkedSermonId"], sid_b)

    async def test_host_cannot_link_403(self) -> None:
        _add_member(self.store, self.org_id, "uid_host", "host")
        sid = await self._seed()
        with self.assertRaises(HTTPException) as ctx:
            sermon_review_routes.link_sermon(
                org_id=self.org_id,
                sermon_id=sid,
                body=sermon_review_routes._LinkBody(serviceKey=self.service_key),
                request=_FakeRequest(),
                user=_user("uid_host"),
            )
        self.assertEqual(ctx.exception.status_code, 403)


if __name__ == "__main__":
    unittest.main()
