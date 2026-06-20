# Design Ref: §2.2 Data Flow (Export) + §10.4 — canonical COLUMNS shared with
# validation. In-memory openpyxl write; never touches the filesystem.

from __future__ import annotations

import re
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .models import Sermon
from .validation import COLUMNS

_PROTECTED_COLUMNS: frozenset[str] = frozenset({
    "Sermon ID",
    "Segment ID",
    "Segment Order",
    "Original Text",
    "App Translation",
})

_HEADER_FILL = PatternFill(
    start_color="FF1F2937", end_color="FF1F2937", fill_type="solid"
)
_HEADER_FONT = Font(color="FFFFFFFF", bold=True)
_PROTECTED_FILL = PatternFill(
    start_color="FFF3F4F6", end_color="FFF3F4F6", fill_type="solid"
)
_EDITABLE_FILL = PatternFill(
    start_color="FFEFF6FF", end_color="FFEFF6FF", fill_type="solid"
)

_WRAP = Alignment(wrap_text=True, vertical="top")

# Excel workbooks use XML internally. These control characters are invalid in
# XML 1.0 and cause openpyxl to reject the cell value. Preserve valid
# whitespace such as tabs, newlines, and carriage returns.
_ILLEGAL_XML_CONTROL_CHARACTERS = re.compile(
    r"[\x00-\x08\x0B\x0C\x0E-\x1F]"
)

_DEFAULT_COLUMN_WIDTH: dict[str, int] = {
    "Sermon ID": 18,
    "Segment ID": 10,
    "Segment Order": 14,
    "Original Text": 50,
    "App Translation": 50,
    "Reviewed Translation": 50,
    "Notes": 30,
    "Status": 12,
}


def _xlsx_value(value: object) -> object:
    if isinstance(value, str):
        return _ILLEGAL_XML_CONTROL_CHARACTERS.sub("", value)
    return value


def build_xlsx(sermon: Sermon) -> bytes:
    """Render a Sermon to a Review File .xlsx as bytes.

    The output is deterministic for a fixed sermon, so an unmodified
    round-trip (export → read) is byte-stable for the data fields (cell
    styling may vary across openpyxl versions but data is identical).
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Sermon Review"

    # Header row
    for col_idx, name in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _WRAP
        ws.column_dimensions[get_column_letter(col_idx)].width = (
            _DEFAULT_COLUMN_WIDTH.get(name, 20)
        )

    for row_idx, segment in enumerate(sermon.segments, start=2):
        values: dict[str, object] = {
            "Sermon ID": sermon.sermonId,
            "Segment ID": segment.segmentId,
            "Segment Order": segment.order,
            "Original Text": segment.original,
            "App Translation": segment.appTranslation,
            "Reviewed Translation": segment.reviewedTranslation,
            "Notes": segment.notes,
            "Status": segment.status,
        }
        for col_idx, name in enumerate(COLUMNS, start=1):
            cell = ws.cell(
                row=row_idx,
                column=col_idx,
                value=_xlsx_value(values[name]),
            )
            cell.alignment = _WRAP
            cell.fill = (
                _PROTECTED_FILL if name in _PROTECTED_COLUMNS else _EDITABLE_FILL
            )

    ws.freeze_panes = "A2"

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
