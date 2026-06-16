# Design Ref: §7 Security (defusedxml-backed, size-capped) + §2.2 Import flow.
# Pure parsing: returns rows + headers. No sermon comparison happens here —
# that lives in validation.py.

from __future__ import annotations

import os
from io import BytesIO

from openpyxl import load_workbook

# Hard cap on rows to defend against pathological inputs even when the file
# is under the size limit. Aligned with SERMON_MAX_SEGMENTS in Design §10.3.
_MAX_ROWS: int = int(os.getenv("SERMON_XLSX_MAX_ROWS", "5000"))


class ImportReadError(ValueError):
    """Raised when a workbook cannot be opened or has structural issues
    that prevent any row-level parsing (e.g. not a zip, no worksheets,
    too many rows)."""


def read_workbook(
    data: bytes,
) -> tuple[list[dict[str, str]], list[str]]:
    """Read an .xlsx and return (rows, headers).

    `rows` is a list of dicts keyed by the header strings (preserving header
    order via dict insertion order). `headers` is the raw header row,
    including any unknown columns the user may have added.
    """
    if not data:
        raise ImportReadError("Empty upload.")

    try:
        wb = load_workbook(
            filename=BytesIO(data), read_only=True, data_only=True
        )
    except Exception as exc:
        raise ImportReadError(f"Could not open workbook: {exc}") from exc

    try:
        ws = wb.active
        if ws is None:
            raise ImportReadError("Workbook has no active worksheet.")

        row_iter = ws.iter_rows(values_only=True)
        try:
            header_tuple = next(row_iter)
        except StopIteration:
            return [], []

        headers = [
            str(h).strip() if h is not None else "" for h in header_tuple
        ]

        rows: list[dict[str, str]] = []
        for r_idx, raw in enumerate(row_iter, start=2):
            if r_idx - 1 > _MAX_ROWS:
                raise ImportReadError(
                    f"Workbook exceeds maximum row count ({_MAX_ROWS})."
                )
            if _is_blank_row(raw):
                continue
            row_dict: dict[str, str] = {}
            for header, value in zip(headers, raw):
                if not header:
                    continue
                row_dict[header] = _coerce(value)
            rows.append(row_dict)
        return rows, headers
    finally:
        wb.close()


def _is_blank_row(values: tuple) -> bool:
    return all(v is None or (isinstance(v, str) and not v.strip()) for v in values)


def _coerce(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value
    return str(value)
