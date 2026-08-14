#!/usr/bin/env python3
"""Lint a sermon manuscript for repeated-phrase patterns that trip PMM.

Two input modes:

    # From an .xlsx sermon-review file
    python -m scripts.sermon_lint --file "path/to/sermon.xlsx"

    # From a live Firestore sermon (requires FIREBASE_ADMIN_CREDENTIALS)
    python -m scripts.sermon_lint --org ark --sermon srm_APrh0qQ_ElUKLgR4

Exit code: 0 if no high-severity collisions, 1 if any high-severity
(gap>=30). Useful in a pre-Sunday CI check.
"""

from __future__ import annotations

import argparse
import json
import os
import sys
from pathlib import Path
from typing import Any

# Allow running as `python scripts/sermon_lint.py` from backend/ as well as
# `python -m scripts.sermon_lint` — normalize sys.path either way.
_SCRIPT_DIR = Path(__file__).resolve().parent
_BACKEND_DIR = _SCRIPT_DIR.parent
if str(_BACKEND_DIR) not in sys.path:
    sys.path.insert(0, str(_BACKEND_DIR))

from app.sermon_review.lint import lint_sermon_segments  # noqa: E402


def _load_from_xlsx(path: str) -> list[dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise SystemExit(
            "openpyxl not installed. Run inside the backend venv: "
            "backend/.venv/bin/python -m scripts.sermon_lint ..."
        ) from exc

    wb = load_workbook(path, read_only=True, data_only=True)
    # Prefer the well-known sheet name from xlsx_export, fall back to first sheet.
    sheet_names = wb.sheetnames
    sheet_name = "Sermon Review" if "Sermon Review" in sheet_names else sheet_names[0]
    ws = wb[sheet_name]

    segments: list[dict[str, Any]] = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue  # header
        row = row + (None, None, None)  # pad
        sid, ko, _en = row[:3]
        if not sid:
            continue
        segments.append({"segmentId": str(sid).strip(), "original": (ko or "").strip()})
    return segments


def _load_from_firestore(org_id: str, sermon_id: str) -> list[dict[str, Any]]:
    from app.services.multichurch_store import multichurch_store

    sermon = multichurch_store.get_review_sermon(org_id, sermon_id)
    if sermon is None:
        raise SystemExit(f"Sermon not found: org={org_id} sermon={sermon_id}")
    return sermon.get("segments") or []


def _print_report(report: dict[str, Any]) -> None:
    total = report["totalSegments"]
    collisions = report["collisions"]
    print(f"Sermon: {total} segments scanned.")
    if not collisions:
        print("No repeated-phrase collisions detected. PMM should track cleanly.")
        return
    print(f"Found {len(collisions)} collision pair(s):\n")
    for c in collisions:
        marker = "!" if c["severity"] == "high" else "·"
        print(
            f"  {marker} {c['shorterSegmentId']} ⊂ {c['longerSegmentId']}  "
            f"(gap={c['gap']}, severity={c['severity']})"
        )
        print(f"      matched text: {c['matchedText']}")
    high_count = sum(1 for c in collisions if c["severity"] == "high")
    if high_count:
        print(
            f"\n{high_count} high-severity collision(s) (gap>=30). "
            f"With the current cursor-sync (nearest-forward) these are safe, "
            f"but you may want to review the manuscript for intentional "
            f"differentiation."
        )


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    src = parser.add_mutually_exclusive_group(required=True)
    src.add_argument("--file", help="Path to a .xlsx sermon-review file")
    src.add_argument(
        "--sermon",
        help="Firestore sermonId (must also pass --org). Requires "
        "FIREBASE_ADMIN_CREDENTIALS.",
    )
    parser.add_argument("--org", help="Firestore orgId (paired with --sermon)")
    parser.add_argument(
        "--json", action="store_true", help="Emit JSON instead of a human report"
    )
    args = parser.parse_args()

    if args.file:
        segments = _load_from_xlsx(args.file)
    else:
        if not args.org:
            parser.error("--sermon requires --org")
        segments = _load_from_firestore(args.org, args.sermon)

    report = lint_sermon_segments(segments)
    if args.json:
        json.dump(report, sys.stdout, ensure_ascii=False, indent=2)
        sys.stdout.write("\n")
    else:
        _print_report(report)

    high_count = sum(1 for c in report["collisions"] if c["severity"] == "high")
    return 1 if high_count > 0 else 0


if __name__ == "__main__":
    raise SystemExit(main())
